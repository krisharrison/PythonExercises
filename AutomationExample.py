import openpyxl as xl #using an alias
from openpyxl.chart import BarChart, Reference


def proccess_workbook(filename):
    wb = xl.load_workbook(filename) #load a numbers wordbook
    sheet = wb['Sheet1']#returns word book object. File has one sheet. Returns 1st sheet

    #access a cell
    #cell = sheet['a1']#coordinates of a sheet using the columns. Column a, row 1.

    #second way to access a cell using sheet object
    #cell = sheet.cell(1, 1) # (row, column)

    #add column and enter new values
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3) #iterate through row of column 3, object
        corrected_price = cell.value * 0.9 # multiply each row by .90. store in corrected_price
        corrected_price_cell = sheet.cell(row,4) # object of rows 2 to 4, column 4
        corrected_price_cell.value = corrected_price #set values of rows 2 to 4, column 4 to corrected_price results

    print("Values inserted")

    #Storing all values in the Reference object
    values = Reference(sheet, min_row=2,max_row=sheet.max_row, min_col=4, max_col = 4)#Max and Min rows...2 - 4

    print("Chart created")
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')



    wb.save(filename) #Save everything to new file
    print("File created")